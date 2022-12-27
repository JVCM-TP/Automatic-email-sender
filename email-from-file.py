"""
A module for reading a specific column from an Excel file.

This module provides a function called `read_excel_column()` that takes an Excel file path, sheet name, and column name as input and returns a list of the values in that column.
"""

import openpyxl

def read_excel_column(file_path: str, sheet_name: str, column_name: str) -> list:
    """
    Read a specific column from an Excel file and return the values as a list.

    :param file_path: The path to the Excel file.
    :param sheet_name: The name of the sheet to read from.
    :param column_name: The name of the column to read.
    :return: A list of the values in the specified column.
    """
    # Open the Excel file
    wb = openpyxl.load_workbook(file_path)

    # Get the sheet you want to read from
    sheet = wb.get_sheet_by_name(sheet_name)

    # Get the column index from the column name
    column_index = sheet.iter_cols(min_col=1, max_col=sheet.max_column, max_row=1).index(column_name) + 1

    # Create an empty array to store the values
    values = []

    # Iterate over the rows in the sheet
    for row in sheet.iter_rows():
        # Get the value of the cell in the specified column
        value = row[column_index].value

        # Add the value to the array
        values.append(value)

    return values
